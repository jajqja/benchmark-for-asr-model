"""Pha 4 — gộp leaderboard + error-analysis từng model thành 1 file .xlsx.

Sheet "Leaderboard": bảng xếp hạng (từ leaderboard.csv).
Mỗi model 1 sheet: utterance WER cao nhất + cặp thay thế + từ hay nuốt/thêm.

Ví dụ:
  PYTHONPATH=src python -m benchmark.report.to_xlsx \
    --csv results/reports/leaderboard.csv \
    --hyp-glob 'results/hypotheses/*/*.jsonl' \
    --out results/reports/report.xlsx
"""
from __future__ import annotations

import argparse
import csv
import glob
import os
import sys

_SRC = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
if _SRC not in sys.path:
    sys.path.insert(0, _SRC)

from benchmark.report.error_analysis import analyze  # noqa: E402


def _load_yaml(path):
    import yaml
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _sheet_name(name, used):
    # Excel: <=31 ký tự, không chứa [ ] : * ? / \
    safe = "".join(c for c in name if c not in "[]:*?/\\")[:31] or "sheet"
    base, i = safe, 1
    while safe in used:
        suffix = f"~{i}"
        safe = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(safe)
    return safe


def _style_header(ws, row, ncols, fill="1F4E78"):
    from openpyxl.styles import Font, PatternFill, Alignment
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(vertical="center")


def _section(ws, r, title):
    from openpyxl.styles import Font
    ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=12, color="1F4E78")
    return r + 1


def build_leaderboard_sheet(ws, csv_path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    if rows:
        _style_header(ws, 1, len(rows[0]))
        ws.freeze_panes = "A2"
        for c in range(1, len(rows[0]) + 1):
            ws.column_dimensions[chr(64 + c)].width = 15 if c > 2 else 24


def build_model_sheet(ws, res, model):
    r = _section(ws, 1, f"{model} — {res['num_utterances']} utterance")
    r += 1

    r = _section(ws, r, "① Utterance WER cao nhất")
    hdr = ["WER%", "#từ ref", "utt_id", "reference", "hypothesis"]
    for c, h in enumerate(hdr, 1):
        ws.cell(row=r, column=c, value=h)
    _style_header(ws, r, len(hdr))
    r += 1
    for u in res["worst"]:
        ws.cell(row=r, column=1, value=round(u["wer"] * 100, 1))
        ws.cell(row=r, column=2, value=u["ref_words"])
        ws.cell(row=r, column=3, value=u["utt_id"])
        ws.cell(row=r, column=4, value=u["ref"])
        ws.cell(row=r, column=5, value=u["hyp"])
        r += 1
    r += 1

    r = _section(ws, r, "② Cặp thay thế hay gặp (ref → hyp)")
    for c, h in enumerate(["ref", "hyp", "count"], 1):
        ws.cell(row=r, column=c, value=h)
    _style_header(ws, r, 3, fill="C55A11")
    r += 1
    for s in res["top_substitutions"]:
        ws.cell(row=r, column=1, value=s["ref"])
        ws.cell(row=r, column=2, value=s["hyp"])
        ws.cell(row=r, column=3, value=s["count"])
        r += 1
    r += 1

    r = _section(ws, r, "③ Từ hay bị NUỐT (deletion)")
    for c, h in enumerate(["từ", "count"], 1):
        ws.cell(row=r, column=c, value=h)
    _style_header(ws, r, 2, fill="548235")
    r += 1
    for d in res["top_deletions"]:
        ws.cell(row=r, column=1, value=d["word"])
        ws.cell(row=r, column=2, value=d["count"])
        r += 1
    r += 1

    r = _section(ws, r, "④ Từ hay bị THÊM (insertion)")
    for c, h in enumerate(["từ", "count"], 1):
        ws.cell(row=r, column=c, value=h)
    _style_header(ws, r, 2, fill="7030A0")
    r += 1
    for i in res["top_insertions"]:
        ws.cell(row=r, column=1, value=i["word"])
        ws.cell(row=r, column=2, value=i["count"])
        r += 1

    for col, w in {"A": 10, "B": 10, "C": 26, "D": 70, "E": 70}.items():
        ws.column_dimensions[col].width = w


def main(argv=None):
    from openpyxl import Workbook

    ap = argparse.ArgumentParser(description="Xuất leaderboard + errors ra .xlsx")
    ap.add_argument("--csv", default="results/reports/leaderboard.csv")
    ap.add_argument("--hyp-glob", default="results/hypotheses/*/*.jsonl")
    ap.add_argument("--manifest-dir", default="data/manifests")
    ap.add_argument("--norm-config", default="configs/normalization/vi.yaml")
    ap.add_argument("--out", default="results/reports/report.xlsx")
    ap.add_argument("--top", type=int, default=25)
    args = ap.parse_args(argv)

    norm_cfg = _load_yaml(args.norm_config)
    wb = Workbook()
    used = set()

    ws = wb.active
    ws.title = _sheet_name("Leaderboard", used)
    build_leaderboard_sheet(ws, args.csv)

    hyps = sorted(p for p in glob.glob(args.hyp_glob) if not p.endswith(".meta.json"))
    n_model = 0
    for hyp in hyps:
        model = os.path.basename(os.path.dirname(hyp))
        ds = os.path.basename(hyp)[:-6]
        manifest = os.path.join(args.manifest_dir, f"{ds}.jsonl")
        if not os.path.exists(manifest):
            print(f"BỎ QUA {hyp} — thiếu {manifest}", file=sys.stderr)
            continue
        res = analyze(manifest, hyp, norm_cfg, top=args.top)
        # nếu nhiều dataset cho cùng model -> đính kèm ds vào tên sheet
        title = model if len(hyps) == len({os.path.basename(os.path.dirname(h)) for h in hyps}) \
            else f"{model}·{ds}"
        ws = wb.create_sheet(_sheet_name(title, used))
        build_model_sheet(ws, res, f"{model} / {ds}")
        n_model += 1
        print(f"  sheet: {ws.title}", file=sys.stderr)

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    wb.save(args.out)
    print(f"-> {args.out}  ({n_model} model + Leaderboard)", file=sys.stderr)


if __name__ == "__main__":
    main()
